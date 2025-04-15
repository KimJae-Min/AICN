import streamlit as st
import zipfile
import os
import torch
import easyocr
import numpy as np
import cv2
from PIL import ImageFont, ImageDraw, Image
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import uuid
import hashlib

st.set_page_config(layout='wide')

@st.cache_resource
def load_model():
    car_m = torch.hub.load("ultralytics/yolov5", 'yolov5s', force_reload=True, skip_validation=True)
    lp_m = torch.hub.load('ultralytics/yolov5', 'custom', 'lp_det.pt')
    reader = easyocr.Reader(['en'], detect_network='craft', recog_network='best_acc',
                            user_network_directory='lp_models/user_network',
                            model_storage_directory='lp_models/models')
    car_m.classes = [2, 3, 5, 7]
    return car_m, lp_m, reader

def save_uploaded_file(directory, file):
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, file.name)
    with open(file_path, 'wb') as f:
        f.write(file.getbuffer())
    return file_path

def extract_zip(zip_path, extract_to='extracted'):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
    return extract_to

def save_to_excel(file_info, file_name):
    output_dir = os.path.join(os.getcwd(), "excel_outputs")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_file_name = f"{file_name}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, unique_file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "차량 인식 결과"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25

    ws.append(["파일명", "파일 시간", "이미지 경로", "이미지 미리보기", "차량 번호"])

    for row_idx, info in enumerate(file_info, start=2):
        ws.cell(row=row_idx, column=1, value=info['파일명'])
        ws.cell(row=row_idx, column=2, value=info['파일 시간'])
        ws.cell(row=row_idx, column=3, value=info['이미지 경로'])

        try:
            img = ExcelImage(info['이미지 경로'])
            with Image.open(info['이미지 경로']) as pil_img:
                width, height = pil_img.size
            target_width = 150
            scale_factor = target_width / width
            target_height = int(height * scale_factor)
            img.width = target_width
            img.height = target_height
            ws.add_image(img, f"D{row_idx}")
            ws.row_dimensions[row_idx].height = target_height * 0.75
        except Exception as e:
            print(f"이미지 삽입 오류 ({info['이미지 경로']}):", e)

        ws.cell(row=row_idx, column=5, value=info['차량 번호'])

    try:
        wb.save(output_path)
        st.success(f"엑셀 파일이 '{output_path}'로 저장되었습니다.")
    except PermissionError:
        st.error("파일 저장 중 문제가 발생했습니다. 엑셀 파일이 열려있거나 권한 문제일 수 있습니다.")

def detect(car_m, lp_m, reader, path):
    fontpath = "SpoqaHanSansNeo-Light.ttf"
    font = ImageFont.truetype(fontpath, 200)
    im = Image.open(path)
    to_draw = np.array(im)
    results = car_m(im)
    locs = results.xyxy[0]
    result_text = []

    if len(locs) == 0:
        return path, ["검출된 차 없음"]

    for idx, item in enumerate(locs):
        x, y, x1, y1 = [int(it.cpu().detach().numpy()) for it in item[:4]]
        car_im = to_draw[y:y1, x:x1, :].copy()
        result = lp_m(Image.fromarray(car_im))

        for rslt in result.xyxy[0]:
            x2, y2, x3, y3 = [int(it.cpu().detach().numpy()) for it in rslt[:4]]
            im_crop = cv2.cvtColor(cv2.resize(to_draw[y+y2:y+y3, x+x2:x+x3], (224, 128)), cv2.COLOR_BGR2GRAY)
            text = reader.recognize(im_crop)[0][1] if reader.recognize(im_crop) else "인식 실패"
            result_text.append(text)

    return path, result_text

def get_file_hash(file_bytes):
    return hashlib.md5(file_bytes).hexdigest()

def main():
    car_m, lp_m, reader = load_model()

    if 'uploaded_hashes' not in st.session_state:
        st.session_state['uploaded_hashes'] = set()

    st.title("\U0001F697 차량 번호판 자동 인식 시스템")
    menu = ['About', '파일 업로드', '결과 확인 및 수정']
    choice = st.sidebar.selectbox('메뉴', menu)

    if choice == 'About':
        st.markdown("<h2 style='font-size:28px;'>\U0001F4CC 개요</h2>", unsafe_allow_html=True)
        st.markdown("""
        이 웹 애플리케이션은 차량 이미지 또는 폴더(zip 파일)를 업로드하면  
        YOLO 모델과 OCR 기술을 활용해 차량 번호판을 자동으로 인식해주는 시스템입니다.
        """)
        st.markdown("---")
        st.markdown("<h2 style='font-size:28px;'>\U0001F6E0 사용 방법</h2>", unsafe_allow_html=True)
        st.markdown("""
        - 왼쪽 사이드바에서 **'파일 업로드'** 메뉴를 선택하세요.  
        - 이미지 또는 폴더를 **ZIP 파일** 형태로 압축하여 **드래그 앤 드롭** 방식으로 업로드합니다.  
        - 인식 결과 확인 및 수정은 왼쪽 메뉴의 **'결과 확인 및 수정'** 탭으로 이동하세요.  
        """)

    elif choice == '파일 업로드':
        st.markdown("### \U0001F4C2 이미지 / ZIP 파일 업로드")
        uploaded_files = st.file_uploader(
            "파일을 이 영역에 드래그 앤 드롭 하세요. (PNG, JPG, JPEG, ZIP 지원, 최대 200MB)",
            type=['png', 'jpg', 'jpeg', 'zip'],
            accept_multiple_files=True,
        )

        file_info = []
        image_dict = {}

        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_ext = uploaded_file.name.split('.')[-1].lower()

                if file_ext == 'zip':
                    zip_path = save_uploaded_file('uploads', uploaded_file)
                    extracted_folder = extract_zip(zip_path)
                    for root, _, files in os.walk(extracted_folder):
                        for file_name in files:
                            if file_name.lower().endswith(('png', 'jpg', 'jpeg')):
                                file_path = os.path.join(root, file_name)
                                try:
                                    with open(file_path, 'rb') as f:
                                        file_bytes = f.read()
                                    file_hash = get_file_hash(file_bytes)
                                except Exception:
                                    continue

                                if file_hash in st.session_state['uploaded_hashes']:
                                    st.toast(f"⚠️ 중복 이미지 무시됨: {file_name}", icon="⚠️")
                                    continue

                                st.session_state['uploaded_hashes'].add(file_hash)
                                result_path, texts = detect(car_m, lp_m, reader, file_path)
                                license_plate = ", ".join(texts)
                                current_time = datetime.now().isoformat().replace(':', "_")
                                file_info.append({
                                    '파일명': file_name,
                                    '파일 시간': current_time,
                                    '이미지 경로': result_path,
                                    '차량 번호': license_plate
                                })
                                image_dict.setdefault(license_plate, []).append(result_path)

                elif file_ext in ['png', 'jpg', 'jpeg']:
                    file_bytes = uploaded_file.getvalue()
                    file_hash = get_file_hash(file_bytes)

                    if file_hash in st.session_state['uploaded_hashes']:
                        st.toast(f"⚠️ 중복 이미지 무시됨: {uploaded_file.name}", icon="⚠️")
                        continue

                    st.session_state['uploaded_hashes'].add(file_hash)
                    file_path = save_uploaded_file('uploads', uploaded_file)
                    result_path, texts = detect(car_m, lp_m, reader, file_path)
                    license_plate = ", ".join(texts)
                    current_time = datetime.now().isoformat().replace(':', "_")
                    file_info.append({
                        '파일명': uploaded_file.name,
                        '파일 시간': current_time,
                        '이미지 경로': result_path,
                        '차량 번호': license_plate
                    })
                    image_dict.setdefault(license_plate, []).append(result_path)

            st.success("✅ 업로드 되었습니다. 결과는 '결과 확인 및 수정' 탭에서 확인하세요.")
            st.session_state['file_info'] = st.session_state.get('file_info', []) + file_info
            for k, v in image_dict.items():
                st.session_state['image_dict'] = st.session_state.get('image_dict', {})
                st.session_state['image_dict'].setdefault(k, []).extend(v)

    elif choice == '결과 확인 및 수정':
        st.subheader('📁 인식 결과 보기 및 수정, 엑셀 저장')

        if 'file_info' not in st.session_state or not st.session_state['file_info']:
            st.info("📂 아직 업로드된 파일이 없습니다. '파일 업로드' 탭에서 먼저 업로드해 주세요.")
        else:
            file_info = st.session_state['file_info']
            image_dict = st.session_state['image_dict']
            image_dict = {k: v for k, v in image_dict.items() if v}
            sorted_plates = sorted(image_dict.keys())

            col1, col2 = st.columns([1, 4])
            with col1:
                selected_plate = st.radio("차량 번호 선택:", sorted_plates)

            with col2:
                if selected_plate and selected_plate in image_dict:
                    for idx, img_path in enumerate(image_dict[selected_plate]):
                        file_name = os.path.basename(img_path).replace("\\", "_").replace("/", "_")
                        current_info = next((info for info in file_info if info['이미지 경로'] == img_path), None)

                        cols = st.columns([2, 1, 1, 1])
                        with cols[0]:
                            st.image(img_path, caption=f"이미지 {idx+1}", use_container_width=True)
                        with cols[1]:
                            st.markdown("**기존 번호**")
                            st.write(current_info['차량 번호'] if current_info else "N/A")
                        with cols[2]:
                            new_plate = st.text_input("수정할 번호", key=f"input_{file_name}_{idx}")
                        with cols[3]:
                            if st.button("수정", key=f"btn_{file_name}_{idx}"):
                                if new_plate:
                                    image_dict[selected_plate].remove(img_path)
                                    if not image_dict[selected_plate]:
                                        del image_dict[selected_plate]
                                    image_dict.setdefault(new_plate, []).append(img_path)
                                    for info in file_info:
                                        if info['이미지 경로'] == img_path:
                                            info['차량 번호'] = new_plate
                                    st.session_state['file_info'] = file_info
                                    st.session_state['image_dict'] = image_dict
                                    st.rerun()

            st.markdown("---")
            file_name = st.text_input("📥 저장할 엑셀 파일명을 입력하세요 (확장자 제외)", "uploaded_images_info")
            if st.button("📄 엑셀로 저장"):
                save_to_excel(file_info, file_name)

if __name__ == '__main__':
    main()
