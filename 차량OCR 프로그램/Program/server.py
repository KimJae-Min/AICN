import os
import shutil
from pathlib import Path
import zipfile
import uuid
import hashlib
from datetime import datetime

import torch
import numpy as np
import cv2
from PIL import Image, ExifTags, UnidentifiedImageError
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from paddleocr import PaddleOCR

# ==============================
# 환경 세팅
# ==============================
# MPS 비활성화 (Mac용 GPU fallback 방지)
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "0"

def disable_mps():
    if hasattr(torch.backends, "mps"):
        torch.backends.mps.is_available = lambda: False
        torch.backends.mps.is_built = lambda: False

disable_mps()

# Streamlit 페이지 설정
st.set_page_config(layout='wide')

# ==============================
# 모델 로드
# ==============================
@st.cache_resource
def load_model():
    device = torch.device("cpu")
    # 차량 감지 모델
    car_m = torch.hub.load('./yolov5', 'yolov5s', source='local').to(device)
    # 번호판 감지 모델 (커스텀)
    lp_m = torch.hub.load('./yolov5', 'custom', path='lp_det.pt', source='local').to(device)
    # OCR 모델
    ocr = PaddleOCR(use_angle_cls=True, lang='korean', use_gpu=False)
    # 차량 클래스만 필터링 (2: car, 3: motorcycle 등)
    car_m.classes = [2, 3, 5, 7]
    return car_m, lp_m, ocr

# ==============================
# 파일 처리 함수
# ==============================
def safe_path(path):
    """특수문자 및 macOS 숨김파일 처리"""
    p = Path(path)
    if "__MACOSX" in p.parts or p.name.startswith("._"):
        return None
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in p.name)
    new_path = p.parent / safe_name
    if new_path != p:
        try:
            shutil.move(p, new_path)
        except Exception:
            return None
    return str(new_path)

def extract_zip(zip_path):
    """ZIP 재귀 추출 및 안전 파일만 반환"""
    extract_to = f"extracted/{uuid.uuid4().hex}"
    os.makedirs(extract_to, exist_ok=True)
    files = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
    for root, _, filenames in os.walk(extract_to):
        for fn in filenames:
            full_path = os.path.join(root, fn)
            safe = safe_path(full_path)
            if not safe:
                continue
            ext = safe.split('.')[-1].lower()
            if ext in ('png','jpg','jpeg'):
                files.append(safe)
            elif ext == 'zip':
                files.extend(extract_zip(safe))
    return files

def save_uploaded_file(directory, uploaded_file):
    os.makedirs(directory, exist_ok=True)
    path = os.path.join(directory, uploaded_file.name)
    with open(path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return path

def get_file_hash(bytes_data):
    return hashlib.md5(bytes_data).hexdigest()

def get_image_date(path):
    """EXIF 정보 기반 촬영일 반환, 없으면 파일 수정일 사용"""
    try:
        img = Image.open(path)
        exif = img._getexif()
        if exif:
            for tag, val in exif.items():
                name = ExifTags.TAGS.get(tag)
                if name in ('DateTimeOriginal', 'DateTime'):
                    return datetime.strptime(val, '%Y:%m:%d %H:%M:%S')
    except Exception:
        pass
    return datetime.fromtimestamp(os.path.getmtime(path))

# ==============================
# OCR 관련 함수
# ==============================
def deskew_plate(plate_img):
    """번호판 기울기 보정"""
    gray = cv2.cvtColor(plate_img, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    lines = cv2.HoughLines(edges, 1, np.pi/180, 100)
    if lines is None:
        return plate_img
    angles = [(theta*180/np.pi - 90) for rho, theta in (l[0] for l in lines) if -45 < theta*180/np.pi-90 < 45]
    if not angles:
        return plate_img
    median_angle = np.median(angles)
    h, w = plate_img.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), median_angle, 1.0)
    rotated = cv2.warpAffine(plate_img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    return rotated

def group_by_chars(ocr_result, y_thresh=10):
    """OCR 결과를 y좌표 기준으로 라인 그룹핑 후 정렬"""
    lines = []
    for box, (text, _) in ocr_result:
        y_center = (box[0][1] + box[2][1]) / 2
        matched = False
        for line in lines:
            if abs(line[0][0] - y_center) < y_thresh:
                line.extend([(y_center, box, ch) for ch in text])
                matched = True
                break
        if not matched:
            lines.append([(y_center, box, ch) for ch in text])
    lines.sort(key=lambda x: x[0][0])
    sorted_texts = []
    for line in lines:
        line.sort(key=lambda x: x[1][0][0])
        sorted_texts.append(''.join([t[2] for t in line]))
    return sorted_texts

def detect(car_m, lp_m, ocr, path):
    """차량+번호판 감지 및 OCR 수행"""
    try:
        im_pil = Image.open(path).convert("RGB")
    except (UnidentifiedImageError, OSError):
        return path, ["인식 불가 (이미지 파일 아님)"]
    
    img = np.array(im_pil)
    result_text = []

    # 차량 감지
    locs = car_m(im_pil).xyxy[0]
    if len(locs) > 0:
        for item in locs:
            x1, y1, x2, y2 = [int(t.cpu().detach().numpy()) for t in item[:4]]
            car_crop = img[y1:y2, x1:x2, :].copy()
            lp_results = lp_m(Image.fromarray(car_crop))
            for lp in lp_results.xyxy[0]:
                lx1, ly1, lx2, ly2 = [int(t.cpu().detach().numpy()) for t in lp[:4]]
                plate_crop = deskew_plate(car_crop[ly1:ly2, lx1:lx2])
                plate_crop = cv2.resize(plate_crop, (224,128))
                gray = cv2.cvtColor(plate_crop, cv2.COLOR_BGR2GRAY)
                ocr_results = ocr.ocr(cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR), cls=True)
                if ocr_results and ocr_results[0]:
                    combined = ''.join(group_by_chars(ocr_results[0])).replace(',', '').replace('，','').replace('-','').replace('－','')
                    result_text.append(combined)

    # 차량 감지 실패 시 전체 이미지 OCR
    if not result_text:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        ocr_full = ocr.ocr(cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR), cls=True)
        if ocr_full and ocr_full[0]:
            result_text = [''.join(group_by_chars(ocr_full[0])).replace(',', '').replace('，','').replace('-','').replace('－','')]
        else:
            result_text = ["인식 실패"]

    return path, result_text

# ==============================
# 엑셀 저장
# ==============================
def save_to_excel(infos, filename):
    os.makedirs('excel_outputs', exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = f"excel_outputs/{filename}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "차량 인식 결과"

    # 헤더
    ws.append(["촬영일", "파일명", "차량 번호", "이미지 미리보기"])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40

    for idx, info in enumerate(infos, start=2):
        ws.cell(idx,1, info['capture_time'])
        ws.cell(idx,2, info['name'])
        ws.cell(idx,3, info['plate'])
        try:
            img = ExcelImage(info['path'])
            img.width = 150
            img.height = int(150*img.height/img.width)
            ws.add_image(img, f"D{idx}")
            ws.row_dimensions[idx].height = img.height * 0.75
        except:
            continue

    wb.save(out_path)
    st.success(f"엑셀 저장 완료: {out_path}")

# ==============================
# Streamlit UI
# ==============================
def main():
    car_m, lp_m, ocr = load_model()
    if 'file_info' not in st.session_state:
        st.session_state['file_info'] = []

    st.title("🚗 차량 번호판 자동 인식 시스템")
    menu = ['📂 파일 업로드', '🔧 번호판 수정 및 결과 확인', 'ℹ️ About']
    choice = st.sidebar.radio('메뉴 선택', menu)

    # -----------------------------
    # 파일 업로드
    # -----------------------------
    if choice == '📂 파일 업로드':
        st.markdown("### 이미지 또는 ZIP 파일 업로드")
        uploaded = st.file_uploader(
            "이미지 또는 ZIP 파일 선택",
            type=['png','jpg','jpeg','zip'],
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
        log_box = st.empty()
        progress_bar = st.progress(0)

        if uploaded:
            infos = []
            for idx, f in enumerate(uploaded, start=1):
                progress_bar.progress(idx / len(uploaded))
                log_box.info(f"📄 처리 중: {f.name}")
                ext = f.name.split('.')[-1].lower()
                saved_path = save_uploaded_file('uploads', f)
                if ext == 'zip':
                    files = extract_zip(saved_path)
                    for fp in files:
                        path, plates = detect(car_m, lp_m, ocr, fp)
                        infos.append({
                            'capture_time': get_image_date(path).strftime('%Y-%m-%d %H:%M:%S'),
                            'name': Path(fp).name,
                            'plate': ", ".join(plates),
                            'path': path
                        })
                else:
                    path, plates = detect(car_m, lp_m, ocr, saved_path)
                    infos.append({
                        'capture_time': get_image_date(path).strftime('%Y-%m-%d %H:%M:%S'),
                        'name': f.name,
                        'plate': ", ".join(plates),
                        'path': path
                    })
            st.session_state['file_info'] += infos
            st.session_state['file_info'].sort(key=lambda x: datetime.strptime(x['capture_time'],'%Y-%m-%d %H:%M:%S'))
            progress_bar.progress(1.0)
            log_box.success("✅ 모든 파일 업로드 및 인식 완료!")

    # -----------------------------
    # 번호판 수정 및 결과 확인
    # -----------------------------
    elif choice == '🔧 번호판 수정 및 결과 확인':
        if not st.session_state['file_info']:
            st.info("📁 먼저 파일을 업로드하세요.")
            return
        file_info = st.session_state['file_info']

        per_page = 10
        total_pages = (len(file_info) + per_page - 1) // per_page
        page = st.number_input("페이지 선택", 1, total_pages, 1)
        current_items = file_info[(page-1)*per_page: page*per_page]

        cols_left, cols_right = st.columns([1.2,1.8])

        with cols_left:
            st.markdown("### 차량 목록")
            for i, info in enumerate(current_items, start=(page-1)*per_page+1):
                if st.button(f"{i}. {info['plate']} ({info['name']})", key=f"btn_{info['path']}"):
                    st.session_state['selected_plate'] = info

        with cols_right:
            st.markdown("### 선택한 차량 상세보기")
            if 'selected_plate' in st.session_state:
                info = st.session_state['selected_plate']
                img_col1, img_col2 = st.columns([1.5,1])
                with img_col1:
                    st.image(info['path'], caption="📷 원본 이미지", use_container_width=True)
                with img_col2:
                    gray_crop = cv2.resize(cv2.cvtColor(cv2.imread(info['path']), cv2.COLOR_BGR2GRAY),(300,150))
                    st.image(gray_crop, caption="🔍 번호판 확대 보기")
                new_plate = st.text_input("번호판 수정", value=info['plate'], key=f"edit_{info['path']}")
                if st.button("💾 수정 저장", key=f"save_{info['path']}"):
                    info['plate'] = new_plate
                    st.success("수정 내용이 저장되었습니다.")

        st.markdown("---")
        fn = st.text_input("엑셀 파일명","vehicles")
        if st.button("📊 엑셀로 내보내기"):
            save_to_excel(file_info, fn)

    # -----------------------------
    # About
    # -----------------------------
    elif choice == 'ℹ️ About':
        st.markdown("### 시스템 개요")
        st.write("""
        이 시스템은 차량 이미지를 입력받아 YOLOv5 + PaddleOCR을 이용해 
        번호판을 자동 인식하고, Excel 파일로 저장할 수 있습니다.
        """)

if __name__ == '__main__':
    main()
