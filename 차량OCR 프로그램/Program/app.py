import os
import zipfile
import uuid
from pathlib import Path
from datetime import datetime
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ExifTags

from my_models import load_models, detect_car_plate

# ==============================
# Streamlit í˜ì´ì§€ ì„¤ì •
# ==============================
st.set_page_config(layout='wide')

# ==============================
# íŒŒì¼ ì²˜ë¦¬ í•¨ìˆ˜
# ==============================
def save_uploaded_file(directory, uploaded_file):
    os.makedirs(directory, exist_ok=True)
    path = os.path.join(directory, uploaded_file.name)
    with open(path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return path

def extract_zip(zip_path):
    extract_to = f"extracted/{uuid.uuid4().hex}"
    os.makedirs(extract_to, exist_ok=True)
    files = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
    for root, _, filenames in os.walk(extract_to):
        for fn in filenames:
            full_path = os.path.join(root, fn)
            if "__MACOSX" in full_path or fn.startswith("._"):
                continue
            if full_path.lower().endswith(('.png', '.jpg', '.jpeg')):
                files.append(full_path)
            elif full_path.lower().endswith('.zip'):
                files += extract_zip(full_path)
    return files

def get_image_date(path):
    try:
        img = Image.open(path)
        exif = img._getexif()
        if exif:
            for tag, val in exif.items():
                name = ExifTags.TAGS.get(tag)
                if name in ('DateTimeOriginal', 'DateTime'):
                    return datetime.strptime(val, '%Y:%m:%d %H:%M:%S')
    except:
        pass
    return datetime.fromtimestamp(os.path.getmtime(path))

def save_to_excel(infos, filename):
    if not infos:
        st.warning("âš ï¸ ë‚´ë³´ë‚¼ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return
    os.makedirs('excel_outputs', exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = f"excel_outputs/{filename}_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ì°¨ëŸ‰ ì¸ì‹ ê²°ê³¼"
    ws.append(["ì´¬ì˜ì¼", "íŒŒì¼ëª…", "ì°¨ëŸ‰ ë²ˆí˜¸", "ì´ë¯¸ì§€ ë¯¸ë¦¬ë³´ê¸°"])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40

    for idx, info in enumerate(infos, start=2):
        ws.cell(idx, 1, info['capture_time'])
        ws.cell(idx, 2, info['name'])
        ws.cell(idx, 3, info['plate'])
        try:
            for p_img in info['plate_imgs']:
                img = ExcelImage(p_img)
                img.width = 150
                img.height = int(150 * img.height / img.width)
                ws.add_image(img, f"D{idx}")
                ws.row_dimensions[idx].height = img.height * 0.75
        except:
            continue
    wb.save(out_path)
    st.success(f"ğŸ“Š ì—‘ì…€ ì €ì¥ ì™„ë£Œ: `{out_path}`")

# ==============================
# Streamlit ëª¨ë¸ ë¡œë“œ
# ==============================
@st.cache_resource
def load_models_cached():
    return load_models()

# ==============================
# Main
# ==============================
def main():
    plate_detector, ocr_bundle = load_models_cached()

    # ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™”
    if 'file_info' not in st.session_state:
        st.session_state['file_info'] = []
    if 'processed_files' not in st.session_state:
        st.session_state['processed_files'] = set()
    if 'current_page' not in st.session_state:
        st.session_state['current_page'] = 1
    if 'upload_key' not in st.session_state:
        st.session_state['upload_key'] = str(uuid.uuid4())

    st.title("ğŸš— ì°¨ëŸ‰ ë²ˆí˜¸íŒ ìë™ ì¸ì‹ ì‹œìŠ¤í…œ")

    # ì‚¬ì´ë“œë°”: ì—…ë¡œë“œ ì´ˆê¸°í™” ë²„íŠ¼
    if st.sidebar.button("ğŸ—‘ ì—…ë¡œë“œ ì´ˆê¸°í™”"):
        st.session_state['file_info'] = []
        st.session_state['processed_files'] = set()
        st.session_state['current_page'] = 1
        st.session_state['upload_key'] = str(uuid.uuid4())
        st.success("ğŸ“‚ ì—…ë¡œë“œëœ ì‚¬ì§„ê³¼ ë°ì´í„°ê°€ ì´ˆê¸°í™”ë˜ì—ˆìŠµë‹ˆë‹¤.")

    menu = ['ğŸ“‚ íŒŒì¼ ì—…ë¡œë“œ', 'ğŸ”§ ë²ˆí˜¸íŒ ìˆ˜ì • ë° ê²°ê³¼ í™•ì¸', 'â„¹ï¸ About']
    choice = st.sidebar.radio('ë©”ë‰´ ì„ íƒ', menu)

    # ==============================
    # ğŸ“‚ íŒŒì¼ ì—…ë¡œë“œ
    # ==============================
    if choice == 'ğŸ“‚ íŒŒì¼ ì—…ë¡œë“œ':
        col_upload, col_list = st.columns([1, 2])
        with col_upload:
            st.markdown("### ğŸ“ ì´ë¯¸ì§€ / ZIP íŒŒì¼ ì—…ë¡œë“œ")
            st.markdown("- ì´ë¯¸ì§€(.png, .jpg, .jpeg) ë˜ëŠ” ZIP(.zip)")

            uploaded = st.file_uploader(
                "íŒŒì¼ ì„ íƒ ğŸ‘‡",
                type=['png', 'jpg', 'jpeg', 'zip'],
                accept_multiple_files=True,
                label_visibility="collapsed",
                key=st.session_state['upload_key']
            )

            log_box = st.empty()
            progress_bar = st.progress(0)
            infos = []
            skipped_files = []

            if uploaded:
                # ì—…ë¡œë“œ ì œí•œ: ìµœëŒ€ 50ê°œ
                remaining = 50 - len(st.session_state['file_info'])
                files_to_add = []
                for f in uploaded:
                    if f.name not in st.session_state['processed_files'] and remaining > 0:
                        files_to_add.append(f)
                    elif f.name not in st.session_state['processed_files']:
                        skipped_files.append(f.name)

                if skipped_files:
                    st.warning(f"âš ï¸ ìµœëŒ€ 50ê°œ ì œí•œ: ë‹¤ìŒ íŒŒì¼ì€ ì²˜ë¦¬ë˜ì§€ ì•ŠìŒ â†’ {', '.join(skipped_files)}")

                for idx, f in enumerate(files_to_add, start=1):
                    progress_bar.progress(idx / len(files_to_add))
                    log_box.info(f"ğŸ“„ {f.name} ì²˜ë¦¬ ì¤‘...")

                    saved_path = save_uploaded_file('uploads', f)
                    files_to_process = [saved_path]
                    if f.name.lower().endswith('.zip'):
                        files_to_process = extract_zip(saved_path)

                    for fp in files_to_process:
                        file_key = Path(fp).name
                        if file_key in st.session_state['processed_files']:
                            log_box.warning(f"âš ï¸ ì´ë¯¸ ì²˜ë¦¬ëœ íŒŒì¼: {file_key} â†’ ê±´ë„ˆëœ€")
                            continue

                        if remaining <= 0:
                            skipped_files.append(file_key)
                            continue

                        st.session_state['processed_files'].add(file_key)
                        remaining -= 1

                        plates, plate_imgs = detect_car_plate(fp, plate_detector, ocr_bundle)
                        infos.append({
                            'capture_time': get_image_date(fp).strftime('%Y-%m-%d %H:%M:%S'),
                            'name': file_key,
                            'plate': ", ".join(plates),
                            'path': fp,
                            'plate_imgs': plate_imgs
                        })

                if skipped_files:
                    st.warning(f"âš ï¸ ì²˜ë¦¬ë˜ì§€ ì•Šì€ íŒŒì¼: {', '.join(skipped_files)}")

                if infos:
                    st.session_state['file_info'] += infos
                    st.session_state['file_info'].sort(
                        key=lambda x: datetime.strptime(x['capture_time'], '%Y-%m-%d %H:%M:%S')
                    )
                progress_bar.progress(1.0)
                log_box.success("âœ… ì—…ë¡œë“œ ë° ì¸ì‹ ì™„ë£Œ!")

        with col_list:
            if st.session_state['file_info']:
                st.markdown("### ğŸ“¸ ì¸ì‹ëœ ì°¨ëŸ‰ ëª©ë¡")
                for info in st.session_state['file_info']:
                    st.write(f"ğŸ“… {info['capture_time']} | ğŸš˜ {info['plate']} | ğŸ–¼ {info['name']}")

    # ==============================
    # ğŸ”§ ë²ˆí˜¸íŒ ìˆ˜ì • ë° ê²°ê³¼ í™•ì¸
    # ==============================
    elif choice == 'ğŸ”§ ë²ˆí˜¸íŒ ìˆ˜ì • ë° ê²°ê³¼ í™•ì¸':
        if not st.session_state['file_info']:
            st.info("ğŸ“ ë¨¼ì € íŒŒì¼ì„ ì—…ë¡œë“œí•˜ì„¸ìš”.")
            return

        file_info = st.session_state['file_info']
        per_page = 10
        total_pages = (len(file_info) + per_page - 1) // per_page
        current_page = st.session_state['current_page']
        start_idx = (current_page - 1) * per_page
        end_idx = start_idx + per_page
        current_items = file_info[start_idx:end_idx]

        left_col, right_col = st.columns([1, 2])
        with left_col:
            st.markdown("### ì°¨ëŸ‰ ëª©ë¡")
            selected_plate_key = st.radio(
                "ì„ íƒí•  ì°¨ëŸ‰ì„ í´ë¦­í•˜ì„¸ìš”",
                options=[f"{i+start_idx+1}. {info['plate']} ({info['name']})" for i, info in enumerate(current_items)],
                key='selected_plate_radio'
            )
            selected_idx = int(selected_plate_key.split('.')[0]) - 1
            info = file_info[selected_idx]

        with right_col:
            st.markdown("### ì„ íƒí•œ ì°¨ëŸ‰ ìƒì„¸ë³´ê¸°")
            orig_img = info['path']
            plate_imgs = info['plate_imgs']

            img_cols = st.columns(2)
            with img_cols[0]:
                st.image(orig_img, caption="ğŸ“· ì›ë³¸ ì´ë¯¸ì§€", use_container_width=True)
            if plate_imgs:
                with img_cols[1]:
                    st.image(plate_imgs[0], caption="ğŸ” ë²ˆí˜¸íŒ í¬ë¡­ 1", use_container_width=True)

            def update_plate(path):
                for item in st.session_state['file_info']:
                    if item['path'] == path:
                        item['plate'] = st.session_state[f"edit_{path}"]

            st.text_input(
                "ë²ˆí˜¸íŒ ìˆ˜ì •",
                value=info['plate'],
                key=f"edit_{info['path']}",
                on_change=update_plate,
                args=(info['path'],)
            )

        # í˜ì´ì§€ë„¤ì´ì…˜ ë²„íŠ¼
        st.markdown("---")
        st.markdown("#### ğŸ“„ í˜ì´ì§€ ì„ íƒ")
        col_prev, col_page, col_next = st.columns([1, 2, 1])
        with col_prev:
            if st.button("â—€ ì´ì „"):
                st.session_state['current_page'] = max(1, st.session_state['current_page'] - 1)
        with col_page:
            st.markdown(f"### {current_page} / {total_pages}", unsafe_allow_html=True)
        with col_next:
            if st.button("ë‹¤ìŒ â–¶"):
                st.session_state['current_page'] = min(total_pages, st.session_state['current_page'] + 1)

        # ì—‘ì…€ ë‚´ë³´ë‚´ê¸°
        st.markdown("---")
        fn = st.text_input("ì—‘ì…€ íŒŒì¼ëª…", "vehicles")
        if st.button("ğŸ“Š ì—‘ì…€ë¡œ ë‚´ë³´ë‚´ê¸°"):
            save_to_excel(file_info, fn)

    # ==============================
    # â„¹ï¸ About
    # ==============================
    elif choice == 'â„¹ï¸ About':
        st.markdown("### ì‹œìŠ¤í…œ ê°œìš”")
        st.write("""
        YOLOv5 + PaddleOCR ê¸°ë°˜ ì°¨ëŸ‰ ë²ˆí˜¸íŒ ì¸ì‹ ì‹œìŠ¤í…œì…ë‹ˆë‹¤.  
        ì´ë¯¸ì§€ë¥¼ ì—…ë¡œë“œí•˜ë©´ ìë™ìœ¼ë¡œ ë²ˆí˜¸íŒì„ ì¸ì‹í•˜ê³ , ê²°ê³¼ë¥¼ Excelë¡œ ë‚´ë³´ë‚¼ ìˆ˜ ìˆìŠµë‹ˆë‹¤.
        """)


if __name__ == '__main__':
    main()
